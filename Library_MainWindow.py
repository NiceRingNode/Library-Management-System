import os,sys,random,sqlite3
from PyQt5.QtWidgets import QApplication,QMainWindow,QLabel,QFileDialog,\
    QTreeWidgetItem,QTableWidgetItem,QMessageBox,QAbstractItemView
from PyQt5.QtCore import Qt,pyqtSlot,QDir,QIODevice,QTextStream,QFile,\
    QTimer
from PyQt5.QtGui import QIcon,QPixmap,QBrush,QPainter,QImage,QColor
from enum import Enum
from PyQt5.QtSql import QSqlDatabase,QSqlQuery
from openpyxl import load_workbook,Workbook
from ui_MainWindow import Ui_MainWindow
from Linklist import *
from Book_Detail_Widget import *
from Qss_File import Qss

class TreeItemType(Enum):
    ItTop = 101
    ItBookClass = 102

class Library_Management(QMainWindow):
    def __init__(self,parent = None):
        """
        :param parent: 父类，默认为None
        """
        super().__init__(parent)
        self.__ui = Ui_MainWindow()
        self.__ui.setupUi(self)
        self.__book_list = linklist() # 图书链表
        self.__cate_dict = dict() # 记录下类别，方便搜索
        self.__cate_dict['书籍类别'] = [] # 点击这个，全部显示
        self.__picture_dict = dict()
        self.__PreReturnRec = [] # 这个是个列表，下标记录的是在还书框的位置编号，内容记录的是对应在借书框那里的位置编号
        self.__PreBorrowRec = -1 # 必须要记录下来当前最新的借书还书的书的序号

        # 搜索相关
        self.__search_place = {} # 值是text或0，text表示现在里面有内容，键不能重复没办法
        self.__search_index = {'ID':0,'name':1,'ISBN':2,'author':3}
        self.__setUI()
        self.__connectDefaultDB()

    def __setUI(self):
        """
        :param: 无
        :return:无
        """
        self.__ui.book_form_search.setAlternatingRowColors(True)  # 交替颜色
        self.__ui.book_form_add.setAlternatingRowColors(True)  # 交替颜色
        self.__ui.book_form_search.setEditTriggers(QAbstractItemView.NoEditTriggers)  # 不允许编辑搜索框
        self.__ui.book_class.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self.setWindowIcon(QIcon(':/icons/icons/appIcon.ico'))
        self.setWindowTitle('图书管理系统')

        self.__timer = QTimer()  # 创建定时器
        self.__timer.setInterval(5000)  # 20s换一次
        self.__timer.timeout.connect(self.update)  # 不用()
        self.__timer.start()
        self.__pictureDictGenerate()

    """-----------------------------------------------------------------"""
    """---------------------------背景动画相关函数------------------------"""

    def __pictureDictGenerate(self):
        """
        :param: 无
        :return: 无
        """
        prefix = ':/images/images/'
        suffixs = ['color.jpg','flower.jpg','sky light.jpg','straw.jpg','tower.jpg']
        for i,suffix in zip(range(len(suffixs)),suffixs):
            self.__picture_dict[i] = os.path.join(prefix,suffix)
        
    def paintEvent(self,event):
        """
        :param event: PyQt5.QtCore.QEvent,标准绘画事件
        :return: 无
        """
        painter = QPainter(self)
        painter.drawRect(self.rect())
        num = random.randint(0,4)
        pixmap = QPixmap(self.__picture_dict[num])
        painter.drawPixmap(self.rect(),pixmap) # 不需要指定background什么的
        super().paintEvent(event)

    """---------------------------背景动画相关函数------------------------"""
    """-----------------------------------------------------------------"""

    """-----------------------------------------------------------------"""
    """---------------------------书籍录入相关函数------------------------"""

    def __createRow(self, row_no: int, flag: str, mode: str, *args):
        """
        :param row_no: int,当前行号
        :param flag: str,创建行的两个表格位置，在book_form_add创建：'add',在book_form_search创建：'search'
        :param mode: str,创建行的两种模式，默认创建：'default',导入创建：'import'
        :param args: tuple,mode='import'时传入的不定长参数
        :return:无
        """
        self.__ui.book_form_add.blockSignals(True)  # 阻止所有信号的传输
        self.__ui.book_form_search.blockSignals(True)

        if args:
            args_list = list(args)

        item_list = list()
        # ID
        itemID = QTableWidgetItem(str(1000 + self.__ui.book_form_add.rowCount()),
                                  CellType.ctID.value)  # 第一个参数就是内容，ctID是指ID单元格
        itemID.setTextAlignment(Qt.AlignHCenter | Qt.AlignVCenter)
        item_list.append(itemID)

        # 书名
        itemName = QTableWidgetItem('无' if mode == 'default' else args_list[0], CellType.ctName.value)
        itemName.setTextAlignment(Qt.AlignHCenter | Qt.AlignVCenter)
        itemName.setForeground(QColor('#008000'))
        item_list.append(itemName)

        # ISBN
        itemISBN = QTableWidgetItem('100-7-1000-1000-0' if mode == 'default' else args_list[1],
                                    CellType.ctISBN.value)  # 第一个参数是内容，第二个参数是类型，类型可以自定义
        itemISBN.setTextAlignment(Qt.AlignHCenter | Qt.AlignVCenter)
        item_list.append(itemISBN)

        # 作者
        itemAuthor = QTableWidgetItem('无' if mode == 'default' else args_list[2],
                                      CellType.ctAuthor.value)  # widgetItem全是str
        itemAuthor.setTextAlignment(Qt.AlignHCenter | Qt.AlignVCenter)
        item_list.append(itemAuthor)

        # 页数
        itemPC = QTableWidgetItem('100' if mode == 'default' else args_list[3], CellType.ctPageCnt.value)
        itemPC.setTextAlignment(Qt.AlignHCenter | Qt.AlignVCenter)
        item_list.append(itemPC)

        # 出版时间
        itemPD = QTableWidgetItem('2020-10-01' if mode == 'default' else args_list[4], CellType.ctPubDate.value)
        itemPD.setTextAlignment(Qt.AlignHCenter | Qt.AlignVCenter)
        item_list.append(itemPD)

        # 收录时间
        itemED = QTableWidgetItem('2020-10-01' if mode == 'default' else args_list[5], CellType.ctEntDate.value)
        itemED.setTextAlignment(Qt.AlignHCenter | Qt.AlignVCenter)
        item_list.append(itemED)

        # 价格
        itemPrice = QTableWidgetItem('100' if mode == 'default' else args_list[6], CellType.ctPrice.value)
        itemPrice.setTextAlignment(Qt.AlignHCenter | Qt.AlignVCenter)
        item_list.append(itemPrice)

        # 类别
        itemCatelog = QTableWidgetItem('计算机类' if mode == 'default' else args_list[7], CellType.ctCatelog.value)
        itemCatelog.setTextAlignment(Qt.AlignHCenter | Qt.AlignVCenter)
        item_list.append(itemCatelog)

        # 借出状态
        itemState = QTableWidgetItem('无借出', CellType.ctState.value)
        itemState.setTextAlignment(Qt.AlignHCenter | Qt.AlignVCenter)
        item_list.append(itemState)

        # 借出次数
        itemBC = QTableWidgetItem('1', CellType.ctBCnt.value)
        itemBC.setTextAlignment(Qt.AlignHCenter | Qt.AlignVCenter)
        item_list.append(itemBC)

        # 余量
        itemBA = QTableWidgetItem('1' if mode == 'default' else args_list[8], CellType.ctBal.value)
        itemBA.setTextAlignment(Qt.AlignHCenter | Qt.AlignVCenter)
        item_list.append(itemBA)

        self.__setAllItem(row_no, flag, item_list)

        self.__ui.book_form_add.blockSignals(False)
        self.__ui.book_form_search.blockSignals(False)

    def __setAllItem(self, row_no: int, flag: str, item_list: list):
        """
        :param row_no: int,当前行号
        :param flag: str,创建行的两个表格位置，在book_form_add创建：'add',在book_form_search创建：'search'
        :param item_list: list[PyQt5.QtWidgets.QTableWidgetItem],将要设置在表格里的项
        :return: 无
        """
        if flag == 'add':
            for i in range(ColNum.colBal.value + 1):
                self.__ui.book_form_add.setItem(row_no, i, item_list[i])
        else:
            for i in range(ColNum.colBal.value + 1):
                self.__ui.book_form_search.setItem(row_no, i, item_list[i])

    def __createRowByNode(self, row_no: int, book, flag: str):  # 用来按类筛选
        """
        :param row_no: int,当前行号
        :param book: node,传入的书籍节点
        :param flag: str,创建行的两个表格位置，在book_form_add创建：'add',在book_form_search创建：'search'
        :return: 无
        """
        self.__ui.book_form_add.blockSignals(True)
        self.__ui.book_form_search.blockSignals(True)

        item_list = list()
        # ID
        itemID = QTableWidgetItem(book.get_ID(), CellType.ctID.value)  # 第一个参数就是内容，ctID是指ID单元格
        itemID.setTextAlignment(Qt.AlignHCenter | Qt.AlignVCenter)
        item_list.append(itemID)

        # 书名
        itemName = QTableWidgetItem(book.get_name(), CellType.ctName.value)
        itemName.setTextAlignment(Qt.AlignHCenter | Qt.AlignVCenter)
        itemName.setForeground(QColor('#008000'))
        item_list.append(itemName)

        # ISBN
        itemISBN = QTableWidgetItem(book.get_ISBN(), CellType.ctISBN.value)  # 第一个参数是内容，第二个参数是类型，类型可以自定义
        itemISBN.setTextAlignment(Qt.AlignHCenter | Qt.AlignVCenter)
        item_list.append(itemISBN)

        # 作者
        itemAuthor = QTableWidgetItem(book.get_author(), CellType.ctAuthor.value)  # widgetItem全是str
        itemAuthor.setTextAlignment(Qt.AlignHCenter | Qt.AlignVCenter)
        item_list.append(itemAuthor)

        # 页数
        itemPC = QTableWidgetItem(book.get_pageCnt(), CellType.ctPageCnt.value)
        itemPC.setTextAlignment(Qt.AlignHCenter | Qt.AlignVCenter)
        item_list.append(itemPC)

        # 出版时间
        itemPD = QTableWidgetItem(book.get_publishedDate(), CellType.ctPubDate.value)
        itemPD.setTextAlignment(Qt.AlignHCenter | Qt.AlignVCenter)
        item_list.append(itemPD)

        # 收录时间
        itemED = QTableWidgetItem(book.get_entryDate(), CellType.ctEntDate.value)
        itemED.setTextAlignment(Qt.AlignHCenter | Qt.AlignVCenter)
        item_list.append(itemED)

        # 价格
        itemPrice = QTableWidgetItem(book.get_price(), CellType.ctPrice.value)
        itemPrice.setTextAlignment(Qt.AlignHCenter | Qt.AlignVCenter)
        item_list.append(itemPrice)

        # 类别
        itemCatelog = QTableWidgetItem(book.get_catelog(), CellType.ctCatelog.value)
        itemCatelog.setTextAlignment(Qt.AlignHCenter | Qt.AlignVCenter)
        item_list.append(itemCatelog)

        # 借出状态
        itemState = QTableWidgetItem(book.get_state(), CellType.ctState.value)
        itemState.setTextAlignment(Qt.AlignHCenter | Qt.AlignVCenter)
        item_list.append(itemState)

        # 借出次数
        itemBC = QTableWidgetItem(book.get_borrowCnt(), CellType.ctBCnt.value)  # 这里要求是int
        itemBC.setTextAlignment(Qt.AlignHCenter | Qt.AlignVCenter)
        item_list.append(itemBC)

        # 余量
        itemBA = QTableWidgetItem(book.get_balance(), CellType.ctBal.value)  # 这里要求是int
        itemBA.setTextAlignment(Qt.AlignHCenter | Qt.AlignVCenter)
        item_list.append(itemBA)

        self.__setAllItem(row_no, flag, item_list)

        self.__ui.book_form_add.blockSignals(False)
        self.__ui.book_form_search.blockSignals(False)

    def __addNewCatelog(self, cur_catelog, cur_row):
        """
        :param cur_catelog: str,当前书籍的类别
        :param cur_row: int,当前行号
        :return: 无
        """
        # 对TreeWidget更改
        flag = False
        if self.__cate_dict.get(cur_catelog) != None:  # 字典的get方法，获取key对应的value
            flag = True
        par_item = self.__ui.book_class.currentItem()
        par_item_2 = self.__ui.book_class_2.currentItem()
        if par_item == None or par_item_2 == None:
            par_item = self.__ui.book_class.topLevelItem(0)  # 获取第一个最高层，我们这里只有一个
            par_item_2 = self.__ui.book_class_2.topLevelItem(0)
        # 判断下是否已经有了
        if flag == False:  # 如果没有就加上，有就算了
            item = QTreeWidgetItem(TreeItemType.ItBookClass.value)  # 节点类型
            item.setText(0, cur_catelog)
            item_2 = QTreeWidgetItem(TreeItemType.ItBookClass.value)  # 节点类型
            item_2.setText(0, cur_catelog)  # item_2 = item就是不行！
            par_item.addChild(item)
            par_item.setExpanded(True)
            par_item_2.addChild(item_2)
            par_item_2.setExpanded(True)
            tmp_list = list([cur_row])
            self.__cate_dict[cur_catelog] = tmp_list
        else:
            self.__cate_dict[cur_catelog].append(cur_row)
        self.__cate_dict['书籍类别'].append(cur_row)
        return

    def __simultaneousItemChanged(self, row: int, col: int, new_content: str) -> None:
        """
        :param row: int,当前行号
        :param col: int,当前列号
        :param new_content: str,用于更新的新内容
        :return: 无
        """
        self.__ui.book_form_search.item(row, col).setText(new_content)

    @pyqtSlot()
    def on_book_add_clicked(self):  # 添加行
        """
        :param: 无
        :return: 无
        """
        cur_row = self.__ui.book_form_add.rowCount()
        tmp_book = book(str(cur_row + 1001), '无', '100-7-1000-1000-0', '无', '100', '2020-10-01', '2020-10-01', '100',
                        '计算机类')
        tmp_booknode = node(tmp_book)
        self.__book_list.insert(tmp_booknode)
        self.__ui.book_form_add.insertRow(cur_row)  # 同步更新
        self.__ui.book_form_search.insertRow(cur_row)
        self.__createRow(cur_row, 'add', 'default')  # 要用行列号指定单元格，但是只需要传行号，列号可以自动获得
        self.__createRow(cur_row, 'search', 'default')
        self.__addNewCatelog('计算机类', cur_row)  # 调用一次就行了，两个表格都会有

    @pyqtSlot()
    def on_book_delete_clicked(self):  # 删除行
        """
        :param: 无
        :return: 无
        """
        cur_row = self.__ui.book_form_add.currentRow()
        if cur_row == -1:  # 表明现在没有选择任一行，使其等于最大那行
            cur_row = self.__ui.book_form_add.rowCount() - 1
        self.__ui.book_form_add.removeRow(cur_row)
        self.__ui.book_form_search.removeRow(cur_row)
        self.__book_list.erase_by_pos(cur_row)
        self.__fixCatelogDict()
        for cur_row in range(self.__ui.book_form_add.rowCount()):
            cur_item = self.__ui.book_form_add.item(cur_row, 0)
            cur_item.setText(str(1001 + cur_row))  # 修正ID

    def __fixCatelogDict(self):
        """
        :param: 无
        :return: 无
        """
        for cur_row in range(self.__ui.book_form_add.rowCount()):
            cur_catelog = self.__ui.book_form_add.item(cur_row, ColNum.colCatelog.value).text()
            self.__cate_dict[cur_catelog].clear()  # 清空所有元素
        self.__cate_dict['书籍类别'].clear()
        for cur_row in range(self.__ui.book_form_add.rowCount()):
            cur_catelog = self.__ui.book_form_add.item(cur_row, ColNum.colCatelog.value).text()
            self.__cate_dict[cur_catelog].append(cur_row)
            self.__cate_dict['书籍类别'].append(cur_row)

    @pyqtSlot(int, int)
    def on_book_form_add_cellChanged(self, row: int, column: int):
        """
        :param row: int,当前行号
        :param column: int,当前列号
        :return: 无
        """
        cur_item = self.__ui.book_form_add.item(row, column)
        if cur_item == None:
            return
        # print(item.text())
        new_content = cur_item.text()
        item_type = cur_item.type()  # 就是那时候指定的几个type
        if column == ColNum.colCatelog.value:
            self.__addNewCatelog(new_content, row)  # 删除旧类是下面，增加新类是这里
        self.__book_list.modify_by_pos(row, item_type, new_content)
        self.__simultaneousItemChanged(row, column, new_content)  # 小幅度改变
        # self.book_list_traverse()
        return

    def __deleteOldCateloginDict(self, old_catelog: str, old_row: int) -> None:
        """
        :param old_catelog: str,旧的类别
        :param old_row: int,旧的行号
        :return: 无
        """
        self.__cate_dict[old_catelog].remove(old_row)

    @pyqtSlot(int, int)
    def on_book_form_add_cellDoubleClicked(self, row: int, column: int) -> None:  # 添加了新的书籍类别，删除旧的
        """
        :param row: int,当前行号
        :param column: int,当前列号
        :return: 无
        """
        cur_item = self.__ui.book_form_add.item(row, column)
        if column != ColNum.colCatelog.value:
            return
        old_content = cur_item.text()
        self.__deleteOldCateloginDict(old_content, row)

    # @pyqtSlot() 不要加
    def on_book_class_2_itemClicked(self, item, column: int) -> None:
        """
        :param item: PyQt5.QtWidgets.QTableWidgetItem,当前表格中选中的项
        :param column: int,当前列号
        :return: 无
        """
        self.__ui.book_form_add.clearContents()
        item_text = item.text(0)
        tmp_list = self.__cate_dict.get(item_text)  # 如果没有这个类，就返回None
        try:
            if tmp_list == None:
                raise KeyError('Error:%s does not exist in categories' % item_text)
        except KeyError as k:
            print(repr(k))
            sys.exit()
        for i, j in zip(tmp_list, range(len(tmp_list))):
            cur_book = self.__book_list.search_by_pos(i).get_data()  # 都是从0开始算的
            self.__createRowByNode(j, cur_book, 'add')  # 一行行打印
            # print('i:',i) 正常的

    def on_book_form_search_cellClicked(self,row,column):
        """
        :param row: int,当前行号
        :param column: int,当前列号
        :return: 无
        """
        if column != ColNum.colName.value:
            return
        cur_book = self.__book_list.search_by_pos(row)
        self.book_popup = Book_Detail(cur_book)
        self.book_popup.setAttribute(Qt.WA_DeleteOnClose)  # 删除时窗口对象也删掉
        book_name = cur_book.get_data().get_name()
        self.book_popup.setWindowTitle(book_name)
        self.book_popup.setWindowFlag(Qt.Window, True)
        self.book_popup.setWindowFlag(Qt.WindowContextHelpButtonHint, False)  # 无帮助
        self.book_popup.show()
        self.book_popup.BRLog_in_mainwindow.connect(self.__set_main_BRLog_table)

    def __set_main_BRLog_table(self):
        """
        :param: 无
        :return: 无
        """
        all_rows = self.__ui.main_BRLog_table.rowCount()
        for i in range(all_rows):
            self.__ui.main_BRLog_table.removeRow(0)  # 这个彻底删除行
        now = self.__book_list.head()
        flag = True
        cur_book_pos_inlist = 0
        borrow_state = True # True代表没有借出，全部归还
        while now != None and now.next != None and flag == True:
            cur_BRLog = now.get_data().get_BRLog()
            for each_log in cur_BRLog: # 如果没有记录，就不会录入表格
                cur_row = self.__ui.main_BRLog_table.rowCount()
                self.__ui.main_BRLog_table.insertRow(cur_row)
                cur_book_name = now.get_data().get_name()
                itemName = QTableWidgetItem(cur_book_name, CellType.ctName.value)
                itemName.setTextAlignment(Qt.AlignHCenter | Qt.AlignVCenter)
                self.__ui.main_BRLog_table.setItem(cur_row,0,itemName)

                if not each_log.get_IsReturn():
                    borrow_state = False
                item_list = generate_BRLog_items(each_log)
                for i in range(ColNum_BRLog.colRealPeriod.value + 1):
                    self.__ui.main_BRLog_table.setItem(cur_row, i + 1, item_list[i])
            if now == self.__book_list.tail():
                flag = False
                break
            now = now.next
            cur_book_pos_inlist += 1
            if borrow_state:
                self.__ui.book_form_add.item(cur_book_pos_inlist, ColNum.colState.value).setText('无借出')
                self.__book_list.modify_by_pos(cur_book_pos_inlist, CellType.ctState.value, '无借出')
            else:
                self.__ui.book_form_add.item(cur_book_pos_inlist, ColNum.colState.value).setText('有借出')
                self.__book_list.modify_by_pos(cur_book_pos_inlist, CellType.ctState.value, '有借出')

    def keyPressEvent(self,event): # 响应键盘输入，delete键等
        """
        :param event: PyQt5.QtCore.QEvent,标准键盘事件
        :return: 无
        """
        print(event.key())
        if event.key() == Qt.Key_Delete:
            self.__ui.book_delete.clicked.emit()
        super().keyPressEvent(event)

    """---------------------------书籍录入相关函数------------------------"""
    """-----------------------------------------------------------------"""

    """------------------------------------------------------------------"""
    """-------------------------文件导入导出相关函数-----------------------"""
    # txt导入，用QFile和QTextStream
    @pyqtSlot()
    def on_txt_in_clicked(self):
        """
        :param: 无
        :return: 无
        """
        cur_path = QDir.currentPath() # 获取系统当前目录
        title = '打开txt文件'
        filt = '文本文件(*.txt)'
        file_name,_ = QFileDialog.getOpenFileName(self,title,cur_path,filt)
        if file_name == "":
            return
        if self.__readTxt(file_name):
            self.__ui.statusbar.showMessage(file_name)
        else:
            QMessageBox.critical(self,'错误','打开文件失败')

    def __readTxt(self,file_name:str) -> bool:
        """
        :param file_name: str,txt文件名
        :return: bool,是否成功
        """
        file_device = QFile(file_name)
        if not file_device.exists():
            return False
        if not file_device.open(QIODevice.ReadOnly | QIODevice.Text):
            return False
        try:
            file_stream = QTextStream(file_device)
            file_stream.setAutoDetectUnicode(True) # 自动检测unicode
            file_stream.setCodec('utf-8') # 不然不能读取汉字
            while not file_stream.atEnd():
                line_str = file_stream.readLine() # 一行行读，读出来就是str
                self.__import2Table(line_str)
        finally:
            file_device.close()
        return True

    def __import2Table(self,line_str:str) -> None:
        """
        :param line_str: str,从self.__readTxt()读入的单行文件str
        :return: 无
        """
        line_list = line_str.split() # 去掉所有的空格
        cur_row = self.__ui.book_form_add.rowCount()
        self.__ui.book_form_add.insertRow(cur_row)  # 同步更新
        self.__ui.book_form_search.insertRow(cur_row)
        tmp_book = book(str(1000 + self.__ui.book_form_add.rowCount()), line_list[1], line_list[2], line_list[3],
                        line_list[4], line_list[5], line_list[6], line_list[7], line_list[8], line_list[9])
        tmp_booknode = node(tmp_book)  # 我可能脑子有病，干嘛不从0开始
        self.__book_list.insert(tmp_booknode)
        self.__createRow(cur_row, 'add', 'import', line_list[1], line_list[2],
                         line_list[3], line_list[4], line_list[5], line_list[6],
                         line_list[7], line_list[8], line_list[9])  # 第一个是行号，不用传进去
        self.__createRow(cur_row, 'search', 'import', line_list[1], line_list[2],
                         line_list[3], line_list[4], line_list[5], line_list[6],
                         line_list[7], line_list[8], line_list[9])
        self.__addNewCatelog(line_list[8], cur_row)
        
    @pyqtSlot()
    def on_txt_out_clicked(self):
        """
        :param: 无
        :return: 无
        """
        cur_path = QDir.currentPath()
        title = '导出txt文件'
        filt = '文本文件(*.txt)'
        file_name,_ = QFileDialog.getSaveFileName(self,title,cur_path,filt) # 在这里已经创建了一个新的文件了
        if file_name == '':
            return
        if self.__writeTxt(file_name):
            self.__ui.statusbar.showMessage(file_name)
        else:
            QMessageBox.critical(self,'错误','导出失败')

    def __writeTxt(self,file_name):
        """
        :param file_name: str,txt文件名
        :return: 无
        """
        file_device = QFile(file_name)
        if not file_device.open(QIODevice.WriteOnly | QIODevice.Text):
            return False
        try:
            file_stream = QTextStream(file_device)
            file_stream.setAutoDetectUnicode(True) # 自动检测Unicode
            file_stream.setCodec('utf-8')
            self.__exportTable(file_stream)
        finally:
            file_device.close()
        return True

    def __exportTable(self,file_stream):
        """
        :param file_stream: PyQt5.QtCore.QTextStream,文件流
        :return: 无
        """
        for i in range(self.__ui.book_form_add.rowCount()):
            text = ' '.join(self.__ui.book_form_add.item(i,j).text() for j in \
                range(self.__ui.book_form_add.columnCount()))
            file_stream << text << '\n'

    @pyqtSlot()
    def on_Excel_in_clicked(self):
        """
        :param: 无
        :return: 无
        """
        cur_path = QDir.currentPath()
        title = '打开excel文件'
        filt = 'excel文件(*.xlsx *.xls)'
        file_name,_ = QFileDialog.getOpenFileName(self,title,cur_path,filt)
        if file_name == '':
            return
        if self.__readExcel(file_name):
            self.__ui.statusbar.showMessage(file_name)
        else:
            QMessageBox.critical(self,'错误','导入失败')
    # xlrd版
    '''
    def __readExcel(self,file_name):
        wb = xlrd.open_workbook(file_name)
        try:
            sheet0 = wb.sheet_by_index(0) # 拿到第一页
            if not wb.sheet_loaded(sheet_name_or_index = 0):
                raise Exception
        except Exception as e:
            print('Error:xlsx file read failed')
            print(repr(e))
            return False
        for i in range(sheet0.nrows): # 这个不是方法
            line_str = ' '.join(map(str,sheet0.row_values(i))) # 获取这一行的所有值组成的列表
            # line_list = list(pd.Series(sheet1.row_values(i)).astype(str))
            # line_str = ' '.join(line_list)
            self.__import2Table(line_str)
        return True
    '''
    def __readExcel(self,file_name:str):
        """
        :param file_name: str,excel文件名，格式为.xlsx
        :return: 无
        """
        wb = load_workbook(file_name)
        try:
            sheet0 = wb.worksheets[0] # 拿到第一页
        except:
            print('xlsx file read failed')
            return False
        for i in range(1,sheet0.max_row): # 这个不是方法，1是因为标题栏也要输出去
            line_list = [str(sheet0.cell(row = i + 1,column = j + 1).value) for j in range(sheet0.max_column)]
            line_str = ' '.join(line_list).replace('00:00:00 ','') # 替换掉
            self.__import2Table(line_str)
        return True

    @pyqtSlot()
    def on_Excel_out_clicked(self):
        """
        :param: 无
        :return: 无
        """
        cur_path = QDir.currentPath()
        title = '保存为excel文件'
        filt = 'excel文件(*.xlsx *.xls)'
        file_name,_ = QFileDialog.getSaveFileName(self,title,cur_path,filt)
        # the file does not need to exist,so this function will automatically 
        # create the target file if it doesn't exist
        if file_name == '':
            return
        if self.__writeExcel(file_name):
            self.__ui.statusbar.showMessage(file_name)
        else:
            QMessageBox.critical(self,'错误','导出失败')
    # pandas版
    '''
    def __writeExcel(self,file_name):
        data = {}
        for i in range(self.__ui.book_form_add.columnCount()):
            text_str = self.__ui.book_form_add.horizontalHeaderItem(i).text()
            data[text_str] = []
        data_keys = data.keys() # 列表
        for i in range(self.__ui.book_form_add.rowCount()):
            for j,k in zip(data_keys,range(len(data_keys))):
                data[j].append(self.__ui.book_form_add.item(i,k).text())
        df = pd.DataFrame(data)
        try:
            df.to_excel(file_name)
        except Exception as e:
            print('Error:xlsx write failed')
            print(expr(e))
            return False
        return True
    '''
    def __writeExcel(self,file_name):
        """
        :param file_name: str,excel文件名，格式为.xlsx
        :return: 无
        """
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = 'Sheet1'
            for i in range(self.__ui.book_form_add.columnCount()):
                text_str = self.__ui.book_form_add.horizontalHeaderItem(i).text()
                ws.cell(row = 1,column = i + 1).value = text_str
            for i in range(self.__ui.book_form_add.rowCount()):
                for j in range(self.__ui.book_form_add.columnCount()):
                    ws.cell(row = i + 2,column = j + 1).value = self.__ui.book_form_add.item(i,j).text()
                    # row = i + 2才对，首先excel要求从1开始，其次上面已经有了标题栏，如果是+1就会覆写
            wb.save(file_name) # 这才是openpyxl正确的打开文件的方式
        except Exception as e:
            print('Error:xlsx write failed')
            print(repr(e))
            return False
        return True

    """-------------------------文件导入导出相关函数-----------------------"""
    """------------------------------------------------------------------"""

    """------------------------------------------------------------------"""
    """--------------------------数据库操作相关函数------------------------"""
    def __connectDefaultDB(self):
        """
        连接默认的数据库
        :param: 无
        :return: 无
        """
        self.__DB = QSqlDatabase.addDatabase('QSQLITE')
        self.__DB.setDatabaseName('Library_Management.db')
        try:
            if not self.__DB.open():  # 如果存在，就连接，如果不存在，就直接创建再连接，但是只有有了这句话才会创建
                raise FileNotFoundError("file can't be created")
        except FileNotFoundError as f:
            QMessageBox.critical(self, '错误', repr(f))
        query = QSqlQuery()
        if not os.path.exists('Library_Management.db'): # 如果没有就创建
            # 新建文件之后要设置新表和各个字段
            new_table1_str = 'create table book_baseinfo(ID varchar primary key,bookname varchar(30),\
                ISBN varchar(30),author varchar(30),pagecnt varchar(10),publishdate varchar(30),\
                entrydate varchar(30),price varchar(10),catelog varchar(20),\
                borrowstate varchar(20),borrowcnt varchar(20),balance varchar(20),\
                absolutepdfAddr varchar(50),introduction varchar(200))'
            # ID为主键
            query.exec_(new_table1_str) # 创建一个表，然后新建这些字段
            new_table2_str = 'create table book_BRinfo(book_name varchar primary key,\
                reader_ID varchar(30),borrowtime varchar(30),borrowplace varchar(30),\
                lending_period varchar(10),duedate varchar(30),isreturn varchar(10),\
                returntime varchar(30),returnplace varchar(30),realborrowlength varchar(20))'
            # book_name书名为主键
            query.exec_(new_table2_str)
        else:
            # 先对第一个表来录入数据
            select_item1 = ['ID','bookname','ISBN','author','pagecnt',
                'publishdate','entrydate','price','catelog','balance']
            t_str1 = ','.join(select_item1)
            select_str1 = 'select '+ t_str1 +' from book_baseinfo'
            item_idxes1 = list()
            item_content1 = list()
            record_rowCnt = -1 # 用来建表
            if query.exec_(select_str1):
                for i in range(len(select_item1)):
                    item_idxes1.append(query.record().indexOf(select_item1[i]))
                while query.next():
                    record_rowCnt += 1
                    for i in range(len(select_item1)):
                        item_content1[i] = query.value(item_idxes1[i])
                    tmp_book = book(item_content1[0],item_content1[1],item_content1[2],item_content1[3],
                        item_content1[4],item_content1[5],item_content1[6],item_content1[7],
                        item_content1[8],item_content1[9])
                    tmp_booknode = node(tmp_book)
                    self.__book_list.insert(tmp_booknode)
                for i in range(record_rowCnt):
                    cur_row = self.__ui.book_form_add.rowCount()
                    cur_book = self.__book_list.search_by_pos(i).get_data()  # 都是从0开始算的
                    self.__createRowByNode(cur_row, cur_book, 'add')  # 一行行打印
                    self.__createRowByNode(cur_row, cur_book, 'search')  # 一行行打印
            # 再对第二个表录入
            select_item2 = ['book_name', 'reader_ID', 'borrowtime', 'borrowplace',
                'lending_period', 'duedate', 'isreturn', 'returntime',
                'returnplace', 'realborrowlength']
            t_str2 = ','.join(select_item2)
            select_str2 = 'select ' + t_str2 + ' from book_BRinfo'
            item_idxes2 = list()
            item_content2 = list()
            if query.exec_(select_str2):
                for i in range(len(select_item2)):
                    item_idxes2.append(query.record().indexOf(select_item2[i]))
                record_rowCnt = -1
                while query.next():
                    record_rowCnt += 1
                    for i in range(len(select_item2)):
                        item_content2[i] = query.value(item_idxes2[i])
                    book_name = item_content2[0]
                    # 用书名来找到位置，把借还日志插入到链表里面
                    cur_pos = -1
                    for i in range(self.__ui.book_form_search.rowCount()):
                        text = self.__ui.book_form_search.item(i,ColNum.colName.value)
                        if text == book_name:
                            cur_pos = i
                            break
                    cur_statelog = state_log(item_content2[1],item_content2[2],
                        item_content2[3],item_content2[4],item_content2[7],
                        item_content2[8])
                    if item_content2[6] == '已归还':
                        cur_statelog.update_IsReturn()
                    cur_statelog.update_real_lendingperiod(item_content2[9])
                    self.__book_list.modify_by_pos(cur_pos,CellType.ctBRLog.value,cur_statelog)

    def on_db_save_clicked(self):
        """
        保存数据到数据库
        :param: 无
        :return: 无
        """
        self.__DB = QSqlDatabase.addDatabase('QSQLITE')
        self.__DB.setDatabaseName('Library_Management.db')
        self.__DB.open()
        query = QSqlQuery()
        # 先清空所有记录
        #truncate_statement = 'TRUNCATE TABLE book_baseinfo'
        #query.exec_(truncate_statement)
        now = self.__book_list.head()
        flag = True
        while now != None and now.next != None and flag == True:
            insert_statement = 'insert into book_baseinfo values(?,?,?,?,?,?,?,?,?,?)'
            query.prepare(insert_statement)
            query.addBindValue(str(now.get_data().get_ID()))
            query.addBindValue(str(now.get_data().get_name()))
            query.addBindValue(str(now.get_data().get_ISBN()))
            query.addBindValue(str(now.get_data().get_author()))
            query.addBindValue(str(now.get_data().get_pageCnt()))
            query.addBindValue(str(now.get_data().get_publishedDate()))
            query.addBindValue(str(now.get_data().get_entryDate()))
            query.addBindValue(str(now.get_data().get_price()))
            query.addBindValue(str(now.get_data().get_catelog()))
            query.addBindValue(str(now.get_data().get_balance()))
            # 上面是书籍基本数据，现在要把所有的借还书记录搞出来
            if not query.exec_():
                print(query.lastError())
            if now == self.__book_list.tail():
                flag = False
                break
            now = now.next
        insert_statement = 'insert into book_BRinfo values(?,?,?,?,?,?,?,?,?,?)'
        cur_row_cnt = self.__ui.main_BRLog_table.rowCount()
        cur_col_cnt = self.__ui.main_BRLog_table.columnCount()
        if cur_row_cnt == 0:
            return
        for i in range(cur_row_cnt):
            query.prepare(insert_statement)
            for j in range(cur_col_cnt):
                text = self.__ui.main_BRLog_table.item(i,j)
                query.addBindValue(text)
            try:
                if not query.exec_():
                    raise ValueError(query.lastError())
            except ValueError as v:
                print('Error:', repr(v))

    """--------------------------数据库操作相关函数------------------------"""
    """------------------------------------------------------------------"""

    """------------------------------------------------------------------"""
    """---------------------------书籍搜索相关函数------------------------"""
    # b=list(s.keys())[list(s.values()).index(200)]
    @pyqtSlot()
    def on_EditID_editingFinished(self):
        """
        :param: 无
        :return: 无
        """
        text = self.__ui.EditID.text()
        if self.__search_place.get('ID', 0) == 0:
            self.__search_place['ID'] = text
        return

    @pyqtSlot()
    def on_EditName_editingFinished(self):
        """
        :param: 无
        :return: 无
        """
        text = self.__ui.EditName.text()
        if self.__search_place.get('name', 0) == 0:
            self.__search_place['name'] = text
        return

    @pyqtSlot()
    def on_EditISBN_editingFinished(self):
        """
        :param: 无
        :return: 无
        """
        text = self.__ui.EditISBN.text()
        if self.__search_place.get('ISBN', 0) == 0:
            self.__search_place['ISBN'] = text
        return

    @pyqtSlot()
    def on_EditAuthor_editingFinished(self):
        """
        :param: 无
        :return: 无
        """
        text = self.__ui.EditAuthor.text()
        if self.__search_place.get('author', 0) == 0:
            self.__search_place['author'] = text
        return

    @pyqtSlot()
    def on_EditPrice_editingFinished(self):
        pass

    @pyqtSlot()
    def on_EditPrice2_editingFinished(self):
        pass

    def search_output(self,index_text):
        """
        :param index_text: int,需要找的节点的下标
        :return: 无
        """
        cur_book = self.__book_list.search_by_pos(index_text).get_data()  # 都是从0开始算的
        self.__createRowByNode(0, cur_book, 'search')


    @pyqtSlot()
    def on_start_search_clicked(self):
        """
        :param: 无
        :return: 无
        """
        # search_filled = list(self.__search_place.keys())[list(self.__search_place.values()).index(1)]
        search_filled = list(self.__search_place.keys()) # 获取所有的有东西的key
        # print(search_filled)
        if search_filled == []:
            QMessageBox.critical(self, '错误', '无搜索关键词')
            return
        # remain_limit = (self.__ui.remain_choose.checkedState == Qt.Checked)
        # ID不用搜，直接拿出来
        if 'ID' in search_filled:
            content_text = int(self.__search_place['ID']) - 1001
            if content_text >= self.__book_list.size():
                QMessageBox.critical(self, '错误', '查无此书')
                return
            # 链表从0开始，行列也是从0开始，但是index是自己设置的，是从1开始
            self.__ui.book_form_search.clearContents()
            self.search_output(content_text)
            self.__search_place.clear()
            return
        flag = False
        first_key = search_filled[0]
        content_text = self.__search_place[first_key]
        cur_col = self.__search_index[first_key]
        for cur_row in range(self.__ui.book_form_add.rowCount()):
            cur_item = self.__ui.book_form_add.item(cur_row,cur_col)
            if cur_item.text() == content_text:
                self.__ui.book_form_search.clearContents()
                self.search_output(cur_row)
                flag = True
                break
        if not flag:
            QMessageBox.critical(self, '错误', '查无此书')
        self.__search_place.clear()
        return

    """---------------------------书籍搜索相关函数------------------------"""
    """------------------------------------------------------------------"""

    def book_list_traverse(self):
        """
        :param: 无
        :return: 无
        """
        now = self.__book_list.head()
        flag = True
        while now != None and now.next != None and flag == True:
            print(now.get_data().get_ID(),now.get_data().get_catelog(),now.get_data().get_name())
            if now == self.__book_list.tail():
                flag = False
                break
            now = now.next

def main():
    app = QApplication(sys.argv)
    app_style = Qss()
    '''
    qss_file = QFile('AppStyle.qss')
    # qss_file = QFile('AppStyle1.qss')
    qss_file.open(QFile.ReadOnly)
    qt_bytes = qss_file.readAll() # QByteArray
    py_bytes = qt_bytes.data() # 转为python的bytes
    app_style = py_bytes.decode('utf-8')
    app.setStyleSheet(app_style)
    '''
    app.setStyleSheet(app_style)
    form = Library_Management()
    form.show()
    sys.exit(app.exec_())

if __name__ == '__main__':
    main()